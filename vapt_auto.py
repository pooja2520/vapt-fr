import platform
import subprocess
import socket
import re
import requests
from requests.auth import HTTPBasicAuth
import json
from datetime import datetime
from urllib.parse import urljoin, urlparse, parse_qs, urlencode
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import warnings
from bs4 import BeautifulSoup
from collections import deque
import xml.etree.ElementTree as ET
warnings.filterwarnings('ignore')

# ==================== UTILITY FUNCTIONS ====================

def run_command(command, timeout=300):
    """Execute shell command safely"""
    try:
        result = subprocess.check_output(
            command, shell=True, stderr=subprocess.STDOUT, timeout=timeout
        )
        return result.decode(errors="ignore")
    except subprocess.CalledProcessError as e:
        return e.output.decode(errors="ignore")
    except subprocess.TimeoutExpired:
        return "Command timed out"
    except Exception as e:
        return f"Error: {str(e)}"

def create_authenticated_session(auth_credentials):
    """Create and configure an authenticated session"""
    session = requests.Session()
    session.verify = False
    
    if not auth_credentials:
        return session
    
    auth_type = auth_credentials.get('type')
    auth_data = auth_credentials.get('data', {})
    session_info = auth_credentials.get('session')
    
    if auth_type == 'basic':
        # HTTP Basic Authentication
        username = auth_data.get('username')
        password = auth_data.get('password')
        if username and password:
            session.auth = (username, password)
            print(f"[*] Using HTTP Basic Authentication for session")
    
    elif auth_type == 'form' and session_info:
        # Form-based authentication - restore cookies from successful login
        stored_cookies = session_info.get('cookies', {})
        if stored_cookies:
            for name, value in stored_cookies.items():
                session.cookies.set(name, value)
            print(f"[*] Using Form Authentication cookies: {list(stored_cookies.keys())}")
        else:
            print("[!] Warning: Form auth selected but no cookies found")
    
    elif auth_type == 'bearer':
        # Bearer token authentication
        token = auth_data.get('token')
        if token:
            session.headers['Authorization'] = f'Bearer {token}'
            print(f"[*] Using Bearer Token Authentication")
    
    elif auth_type == 'custom':
        # Custom headers
        custom_headers = auth_data.get('headers', {})
        if custom_headers:
            session.headers.update(custom_headers)
            print(f"[*] Using Custom Headers Authentication")
    
    return session

def resolve_target(target):
    """Resolve domain to IP address"""
    try:
        if not target.startswith(('http://', 'https://')):
            domain = target
        else:
            domain = urlparse(target).netloc or target
        
        ip = socket.gethostbyname(domain)
        return ip
    except:
        return "Unable to resolve"

def get_base_url(target):
    """Get proper URL format"""
    if not target.startswith(('http://', 'https://')):
        return f"http://{target}"
    return target

def get_domain(target):
    """Extract domain from URL"""
    if target.startswith(('http://', 'https://')):
        return urlparse(target).netloc
    return target

def check_web_application(target):
    """Check if target is a web application"""
    print("\n[+] Checking if target is a web application...")
    
    try:
        url = get_base_url(target)
        response = requests.get(url, timeout=10, allow_redirects=True, verify=False)
        
        if response.status_code in range(200, 600):
            content_type = response.headers.get('Content-Type', '').lower()
            is_web = (
                'text/html' in content_type or
                'application/json' in content_type or
                'application/xml' in content_type or
                '<html' in response.text.lower() or
                '<!doctype html' in response.text.lower()
            )
            
            if is_web:
                print(f"[✓] Web application detected on {url}")
                return True
            else:
                print(f"[!] Target is not a web application")
                return False
        else:
            return False
            
    except Exception as e:
        print(f"[!] Cannot reach target: {str(e)}")
        return False

# ==================== NETWORK TESTS ====================

def test_reconnaissance(target):
    """Perform reconnaissance"""
    print("\n[+] Performing Reconnaissance...")
    results = []
    
    domain = get_domain(target)
    ip = resolve_target(domain)
    
    results.append({
        'Test': 'DNS Resolution',
        'Finding': f'Target: {target}, Resolved IP: {ip}',
        'Severity': 'Info',
        'Status': 'Complete',
        'Vulnerable Path': 'N/A - Information Gathering',
        'Remediation': 'N/A',
        'Resolution Steps': 'N/A'
    })
    
    # Ping test
    if platform.system().lower() == "windows":
        ping_cmd = f"ping -n 2 {domain}"
    else:
        ping_cmd = f"ping -c 2 {domain}"
    
    ping_output = run_command(ping_cmd)
    
    if 'bytes=' in ping_output or 'ttl=' in ping_output.lower():
        results.append({
            'Test': 'Ping Test',
            'Finding': 'Host is reachable via ICMP',
            'Severity': 'Info',
            'Status': 'Success',
            'Vulnerable Path': 'ICMP Protocol',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    else:
        results.append({
            'Test': 'Ping Test',
            'Finding': 'Host may be blocking ICMP or unreachable',
            'Severity': 'Low',
            'Status': 'Failed',
            'Vulnerable Path': 'Network Configuration',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_port_scanning(target):
    """Perform port scanning"""
    print("\n[+] Port Scanning...")
    results = []
    
    domain = get_domain(target)
    cmd = f"nmap -Pn -T4 --top-ports 100 {domain}"
    output = run_command(cmd)
    
    open_ports = []
    for line in output.splitlines():
        match = re.match(r'(\d+)/(\w+)\s+open\s+(.+)', line.strip())
        if match:
            port, protocol, service = match.groups()
            open_ports.append(f"{port}/{protocol} ({service})")
    
    if open_ports:
        results.append({
            'Test': 'Port Scan',
            'Finding': f"Open ports found: {', '.join(open_ports)}",
            'Severity': 'Info',
            'Status': 'Complete',
            'Vulnerable Path': ', '.join(open_ports),
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    else:
        results.append({
            'Test': 'Port Scan',
            'Finding': 'No open ports detected in top 100',
            'Severity': 'Info',
            'Status': 'Complete',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_service_detection(target):
    """Detect services and versions"""
    print("\n[+] Service Detection...")
    results = []
    
    domain = get_domain(target)
    cmd = f"nmap -sV --version-intensity 5 {domain}"
    output = run_command(cmd)
    
    services = []
    for line in output.splitlines():
        match = re.match(r'(\d+)/(\w+)\s+open\s+(\S+)\s+(.*)', line.strip())
        if match:
            port, protocol, service, version = match.groups()
            services.append(f"{port}/{protocol}: {service} {version}")
    
    if services:
        finding = "Services detected: " + "; ".join(services)
        results.append({
            'Test': 'Service Detection',
            'Finding': finding,
            'Severity': 'Info',
            'Status': 'Complete',
            'Vulnerable Path': ', '.join([s.split(':')[0] for s in services]),
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    else:
        results.append({
            'Test': 'Service Detection',
            'Finding': 'No services detected',
            'Severity': 'Info',
            'Status': 'Complete',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_vulnerability_scanning(target):
    """Scan for known vulnerabilities"""
    print("\n[+] Vulnerability Scanning...")
    results = []
    
    domain = get_domain(target)
    cmd = f"nmap --script vuln {domain}"
    output = run_command(cmd, timeout=180)
    
    if 'VULNERABLE' in output.upper() or 'CVE-' in output.upper():
        results.append({
            'Test': 'Vulnerability Scan',
            'Finding': 'Potential vulnerabilities detected by automated scan',
            'Severity': 'High',
            'Status': 'Vulnerable',
            'Vulnerable Path': domain,
            'Remediation': 'Update all software to latest versions and apply security patches immediately.',
            'Resolution Steps': '1. Review nmap vulnerability scan output\n2. Identify affected services and versions\n3. Update to patched versions\n4. Apply security hardening\n5. Re-scan to verify fixes'
        })
    else:
        results.append({
            'Test': 'Vulnerability Scan',
            'Finding': 'No known vulnerabilities detected by automated scan',
            'Severity': 'Info',
            'Status': 'Secure',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

# ==================== WEB CRAWLING ====================


def crawl_website(base_url, auth_credentials=None, max_pages=50, progress_callback=None):
    """Crawl website to discover paths with real-time progress updates and proper authentication"""
    print(f"\n[+] Starting web crawler (max {max_pages} pages)...")
    
    if progress_callback:
        progress_callback({'type': 'crawl_start', 'url': base_url, 'max_pages': max_pages})
    
    visited = set()
    to_visit = deque([base_url])
    paths = set()
    forms = []
    pages_crawled = []
    
    # FIXED: Use the new create_authenticated_session function
    session = create_authenticated_session(auth_credentials)
    
    parsed_base = urlparse(base_url)
    base_domain = parsed_base.netloc
    
    while to_visit and len(visited) < max_pages:
        try:
            current_url = to_visit.popleft()
            
            if current_url in visited:
                continue
            
            parsed_current = urlparse(current_url)
            if parsed_current.netloc != base_domain:
                continue
            
            print(f"  [*] Crawling: {current_url}")
            
            # Send progress update
            if progress_callback:
                progress_callback({
                    'type': 'crawling',
                    'url': current_url,
                    'count': len(visited) + 1,
                    'total': max_pages
                })
            
            try:
                response = session.get(current_url, timeout=10, allow_redirects=True)
                visited.add(current_url)
                
                path = parsed_current.path or '/'
                if parsed_current.query:
                    path += '?' + parsed_current.query
                paths.add(current_url)
                
                pages_crawled.append({
                    'url': current_url,
                    'status': response.status_code,
                    'content_type': response.headers.get('Content-Type', '')
                })
                
                if 'text/html' in response.headers.get('Content-Type', ''):
                    soup = BeautifulSoup(response.content, 'html.parser')
                    
                    # Extract forms
                    for form in soup.find_all('form'):
                        form_data = {
                            'action': urljoin(current_url, form.get('action', '')),
                            'method': form.get('method', 'get').upper(),
                            'inputs': []
                        }
                        
                        for input_tag in form.find_all(['input', 'textarea', 'select']):
                            form_data['inputs'].append({
                                'name': input_tag.get('name', ''),
                                'type': input_tag.get('type', 'text'),
                                'value': input_tag.get('value', '')
                            })
                        
                        forms.append(form_data)
                    
                    # Extract links
                    for link in soup.find_all('a', href=True):
                        absolute_url = urljoin(current_url, link['href'])
                        parsed_link = urlparse(absolute_url)
                        
                        if parsed_link.netloc == base_domain:
                            clean_url = absolute_url.split('#')[0]
                            if clean_url and clean_url not in visited:
                                to_visit.append(clean_url)
                
            except Exception as e:
                print(f"  [!] Error crawling {current_url}: {str(e)}")
                continue
                
        except Exception as e:
            continue
    
    print(f"[+] Crawling complete: {len(visited)} pages, {len(paths)} paths discovered")
    
    if progress_callback:
        progress_callback({
            'type': 'crawl_complete',
            'total_paths': len(paths),
            'pages_crawled': len(visited)
        })
    
    return {
        'paths': list(paths),
        'forms': forms,
        'pages_crawled': pages_crawled
    }


# ==================== OWASP TESTS PER PATH ====================

def test_rate_limiting(session, url, path):
    """Test for rate limiting (OWASP requirement)"""
    results = []
    
    try:
        request_count = 30
        responses = []
        start_time = time.time()
        
        for i in range(request_count):
            try:
                resp = session.get(url, timeout=3, verify=False)
                responses.append(resp.status_code)
            except:
                break
        
        elapsed_time = time.time() - start_time
        rate_limited = 429 in responses
        all_successful = all(status == 200 for status in responses)
        
        if not rate_limited and all_successful and elapsed_time < 5:
            results.append({
                'Test': 'Rate Limiting',
                'Finding': f'No rate limiting detected - {request_count} rapid requests completed without restriction',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement rate limiting to prevent brute force attacks, credential stuffing, and denial of service.',
                'Resolution Steps': '1. Implement rate limiting middleware (e.g., express-rate-limit, Flask-Limiter)\n2. Set limits: 100 requests per 15 minutes per IP for standard endpoints\n3. Set stricter limits for authentication endpoints: 5 failed attempts per 15 minutes\n4. Return 429 (Too Many Requests) when limit exceeded\n5. Use Redis or Memcached for distributed rate limiting\n6. Implement progressive delays for repeated violations\n7. Whitelist trusted IPs if needed\n8. Log all rate limit violations'
            })
        else:
            results.append({
                'Test': 'Rate Limiting',
                'Finding': 'Rate limiting appears to be implemented',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Rate Limiting',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_request_throttling(session, url, path):
    """Test for request throttling"""
    results = []
    
    try:
        # Test with burst requests
        burst_count = 50
        throttled = False
        
        for i in range(burst_count):
            try:
                resp = session.get(url, timeout=2, verify=False)
                if resp.status_code == 429 or 'retry-after' in resp.headers:
                    throttled = True
                    break
            except:
                break
        
        if not throttled:
            results.append({
                'Test': 'Request Throttling',
                'Finding': f'No request throttling detected - burst of {burst_count} requests accepted',
                'Severity': 'Low',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement request throttling to control traffic spikes and prevent resource exhaustion.',
                'Resolution Steps': '1. Implement token bucket or leaky bucket algorithm\n2. Set burst allowance (e.g., 20 requests per second)\n3. Configure sustained rate (e.g., 5 requests per second average)\n4. Return 429 with Retry-After header\n5. Implement per-user and per-IP throttling\n6. Use CDN or reverse proxy for traffic shaping\n7. Monitor and adjust limits based on usage patterns\n8. Implement graceful degradation under load'
            })
        else:
            results.append({
                'Test': 'Request Throttling',
                'Finding': 'Request throttling is implemented',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Request Throttling',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_csp(session, url, path):
    """Test Content Security Policy"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        headers = response.headers
        
        csp_header = headers.get('Content-Security-Policy', '')
        csp_report = headers.get('Content-Security-Policy-Report-Only', '')
        
        if not csp_header and not csp_report:
            results.append({
                'Test': 'Content Security Policy',
                'Finding': 'No Content-Security-Policy header present - vulnerable to XSS and code injection attacks',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement Content Security Policy header to mitigate XSS and data injection attacks.',
                'Resolution Steps': '1. Define CSP policy: Content-Security-Policy: default-src \'self\'\n2. Whitelist specific trusted domains for scripts, styles, images\n3. Use nonces for inline scripts: script-src \'nonce-{random}\'\n4. Avoid unsafe-inline and unsafe-eval\n5. Test with Content-Security-Policy-Report-Only first\n6. Set up CSP violation reporting endpoint\n7. Monitor CSP reports and refine policy\n8. Use strict-dynamic for modern browsers'
            })
        elif 'unsafe-inline' in csp_header.lower() or 'unsafe-eval' in csp_header.lower():
            results.append({
                'Test': 'Content Security Policy',
                'Finding': f'CSP contains unsafe directives: {csp_header[:100]}',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Remove unsafe-inline and unsafe-eval from CSP policy.',
                'Resolution Steps': '1. Audit all inline scripts and styles\n2. Move inline code to external files\n3. Use nonces or hashes for required inline content\n4. Remove unsafe-inline and unsafe-eval directives\n5. Test thoroughly to ensure functionality\n6. Deploy updated CSP policy\n7. Monitor for violations'
            })
        else:
            results.append({
                'Test': 'Content Security Policy',
                'Finding': 'Content Security Policy is properly configured',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Content Security Policy',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_server_disclosure(session, url, path):
    """Test server information disclosure"""
    results = []
    
    try:
        test_urls = [
            url + '/nonexistent-' + str(time.time()),
            url.replace(path, '/error-test-404'),
        ]
        
        verbose_errors = []
        
        for test_url in test_urls[:1]:
            try:
                response = session.get(test_url, timeout=10, verify=False)
                
                error_indicators = [
                    'Warning:', 'Fatal error:', 'Notice:', 'Parse error:',
                    'Exception', 'Stack trace', 'at line', 'Traceback',
                    'SQLSTATE', 'mysql_', 'ORA-', 'SQLException',
                    'PHP Version', 'Server Error', 'Debug', 'Application Error'
                ]
                
                for indicator in error_indicators:
                    if indicator in response.text:
                        verbose_errors.append(indicator)
                        break
                
            except:
                continue
        
        if verbose_errors:
            results.append({
                'Test': 'Server Information Disclosure',
                'Finding': f'Server reveals verbose error messages: {", ".join(set(verbose_errors))}',
                'Severity': 'Low',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Disable verbose error messages and implement custom error pages.',
                'Resolution Steps': '1. Disable display_errors in production (PHP: display_errors = Off)\n2. Configure custom error pages (404, 500, etc.)\n3. Log detailed errors server-side only\n4. Remove stack traces from responses\n5. Set generic error messages for users\n6. Configure web server to hide version information\n7. Implement centralized error logging\n8. Review all error handling code'
            })
        else:
            results.append({
                'Test': 'Server Information Disclosure',
                'Finding': 'No verbose error messages detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Server Information Disclosure',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_hardcoded_secrets(session, url, path):
    """Test for hardcoded secrets"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        content = response.text.lower()
        
        secret_patterns = [
            (r'api[_-]?key["\']?\s*[:=]\s*["\']?([a-zA-Z0-9_-]{20,})', 'API Key'),
            (r'secret[_-]?key["\']?\s*[:=]\s*["\']?([a-zA-Z0-9_-]{20,})', 'Secret Key'),
            (r'password["\']?\s*[:=]\s*["\']([^\s"\']{6,})', 'Password'),
            (r'aws[_-]?access[_-]?key', 'AWS Access Key'),
        ]
        
        found_secrets = []
        
        for pattern, secret_type in secret_patterns:
            if re.search(pattern, content):
                found_secrets.append(secret_type)
        
        if found_secrets:
            results.append({
                'Test': 'Hardcoded Secrets Detection',
                'Finding': f'Potential hardcoded secrets in response: {", ".join(set(found_secrets))}',
                'Severity': 'Critical',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Remove all hardcoded secrets. Use environment variables and secret management.',
                'Resolution Steps': '1. Immediately rotate all exposed credentials\n2. Remove secrets from source code\n3. Use environment variables for configuration\n4. Implement secret management (AWS Secrets Manager, HashiCorp Vault)\n5. Add secrets to .gitignore\n6. Use git-secrets or TruffleHog for scanning\n7. Never commit secrets to version control\n8. Implement pre-commit hooks'
            })
        else:
            results.append({
                'Test': 'Hardcoded Secrets Detection',
                'Finding': 'No obvious hardcoded secrets detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Hardcoded Secrets Detection',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_sql_injection_owasp(session, url, path):
    """Test SQL Injection (OWASP A03:2021)"""
    results = []
    
    try:
        sql_payloads = ["' OR '1'='1", "1' OR '1'='1' --", "1' UNION SELECT NULL--"]
        sql_errors = ['sql syntax', 'mysql', 'sqlite', 'postgresql', 'oracle', 'odbc', 'jdbc', 'SQLSTATE']
        
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            
            for param_name in params.keys():
                for payload in sql_payloads[:1]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, verify=False)
                        for error in sql_errors:
                            if error.lower() in response.text.lower():
                                vulnerable = True
                                break
                        if vulnerable:
                            break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'SQL Injection (OWASP A03:2021)',
                'Finding': 'SQL Injection vulnerability detected - database errors revealed in response',
                'Severity': 'Critical',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Use parameterized queries (prepared statements) for all database operations.',
                'Resolution Steps': '1. Replace all dynamic SQL with parameterized queries\n2. Use ORM frameworks (SQLAlchemy, Hibernate, Entity Framework)\n3. Implement strict input validation\n4. Apply principle of least privilege to database accounts\n5. Use stored procedures where appropriate\n6. Disable detailed error messages\n7. Implement WAF with SQL injection rules\n8. Regular security testing with SQLMap'
            })
        else:
            results.append({
                'Test': 'SQL Injection (OWASP A03:2021)',
                'Finding': 'No SQL injection vulnerability detected in basic testing',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'SQL Injection (OWASP A03:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_command_injection_owasp(session, url, path):
    """Test Command Injection (OWASP A03:2021)"""
    results = []
    
    try:
        cmd_payloads = ['; whoami', '| id', '`pwd`']
        cmd_indicators = ['uid=', 'gid=', 'groups=', 'root:', '/home/', '/bin/']
        
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            
            for param_name in params.keys():
                for payload in cmd_payloads[:1]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, verify=False)
                        for indicator in cmd_indicators:
                            if indicator in response.text:
                                vulnerable = True
                                break
                        if vulnerable:
                            break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'Command Injection (OWASP A03:2021)',
                'Finding': 'Command injection vulnerability detected - OS commands appear to execute',
                'Severity': 'Critical',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Never pass user input to system commands. Use safe APIs instead.',
                'Resolution Steps': '1. Avoid system() and exec() functions with user input\n2. Use language-specific safe APIs (subprocess with shell=False)\n3. Implement strict input validation with whitelisting\n4. Use parameterized APIs\n5. Run application with minimal privileges\n6. Implement command injection filters in WAF\n7. Use security linters and SAST tools\n8. Code review all command execution points'
            })
        else:
            results.append({
                'Test': 'Command Injection (OWASP A03:2021)',
                'Finding': 'No command injection vulnerability detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Command Injection (OWASP A03:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_xss_owasp(session, url, path):
    """Test XSS (OWASP A03:2021)"""
    results = []
    
    try:
        xss_payloads = ['<script>alert(1)</script>', '<img src=x onerror=alert(1)>']
        
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            
            for param_name in params.keys():
                for payload in xss_payloads[:1]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, verify=False)
                        if payload in response.text:
                            vulnerable = True
                            break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'Cross-Site Scripting - XSS (OWASP A03:2021)',
                'Finding': 'XSS vulnerability detected - unescaped user input reflected in response',
                'Severity': 'High',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement proper output encoding and Content Security Policy.',
                'Resolution Steps': '1. Encode all user input before output (HTML, JavaScript, URL, CSS encoding)\n2. Use context-aware encoding libraries\n3. Implement strict Content Security Policy\n4. Validate and sanitize input server-side\n5. Use HTTPOnly and Secure flags on cookies\n6. Implement X-XSS-Protection header\n7. Use frameworks with auto-escaping (React, Angular)\n8. Regular XSS testing and code review'
            })
        else:
            results.append({
                'Test': 'Cross-Site Scripting - XSS (OWASP A03:2021)',
                'Finding': 'No XSS vulnerability detected in basic testing',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Cross-Site Scripting - XSS (OWASP A03:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_csrf_owasp(session, url, path):
    """Test CSRF (OWASP A01:2021)"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        
        csrf_patterns = [r'csrf', r'_token', r'authenticity', r'__requestverification']
        has_csrf = any(re.search(pattern, response.text, re.I) for pattern in csrf_patterns)
        
        set_cookie = response.headers.get('Set-Cookie', '')
        has_samesite = 'samesite=' in set_cookie.lower()
        
        if not has_csrf and not has_samesite:
            results.append({
                'Test': 'CSRF Protection (OWASP A01:2021)',
                'Finding': 'No CSRF protection detected - missing tokens and SameSite cookies',
                'Severity': 'High',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement CSRF tokens for all state-changing operations.',
                'Resolution Steps': '1. Implement CSRF tokens for POST/PUT/DELETE requests\n2. Use synchronizer token pattern\n3. Set SameSite=Strict or Lax on cookies\n4. Validate Origin/Referer headers\n5. Use framework CSRF protection (Django, Spring Security)\n6. Require re-authentication for sensitive operations\n7. Implement custom request headers for AJAX\n8. Never use GET for state changes'
            })
        else:
            results.append({
                'Test': 'CSRF Protection (OWASP A01:2021)',
                'Finding': 'CSRF protection appears to be implemented',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'CSRF Protection (OWASP A01:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_ssrf_owasp(session, url, path):
    """Test SSRF (OWASP A10:2021)"""
    results = []
    
    try:
        ssrf_payloads = ['http://127.0.0.1', 'http://localhost', 'http://169.254.169.254']
        
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            url_params = [p for p in params if any(k in p.lower() for k in ['url', 'uri', 'link', 'redirect', 'src'])]
            
            for param_name in url_params:
                for payload in ssrf_payloads[:1]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, verify=False)
                        if 'localhost' in response.text.lower() or '127.0.0.1' in response.text:
                            vulnerable = True
                            break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'Server-Side Request Forgery - SSRF (OWASP A10:2021)',
                'Finding': 'Potential SSRF vulnerability - server may fetch user-supplied URLs',
                'Severity': 'High',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Validate and restrict URLs. Use allowlist of permitted domains.',
                'Resolution Steps': '1. Implement strict URL validation with allowlist\n2. Disable unused URL schemes (file://, gopher://)\n3. Block private IP ranges (RFC1918)\n4. Block cloud metadata endpoints (169.254.169.254)\n5. Use DNS rebinding protection\n6. Implement network segmentation\n7. Use separate service accounts with minimal permissions\n8. Log and monitor all outbound requests'
            })
        else:
            results.append({
                'Test': 'Server-Side Request Forgery - SSRF (OWASP A10:2021)',
                'Finding': 'No SSRF vulnerability detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Server-Side Request Forgery - SSRF (OWASP A10:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_xxe_owasp(session, url, path):
    """Test XXE (OWASP A05:2021)"""
    results = []
    
    try:
        xxe_payload = '''<?xml version="1.0"?><!DOCTYPE foo [<!ENTITY xxe SYSTEM "file:///etc/passwd">]><root>&xxe;</root>'''
        
        vulnerable = False
        
        try:
            response = session.post(url, data=xxe_payload, headers={'Content-Type': 'application/xml'}, timeout=10, verify=False)
            if 'root:' in response.text or 'daemon:' in response.text:
                vulnerable = True
        except:
            pass
        
        if vulnerable:
            results.append({
                'Test': 'XML External Entity - XXE (OWASP A05:2021)',
                'Finding': 'XXE vulnerability detected - XML parser processes external entities',
                'Severity': 'High',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Disable XML external entity processing in all parsers.',
                'Resolution Steps': '1. Disable DTD processing in XML parsers\n2. Disable external entity expansion\n3. Use less complex formats (JSON instead of XML)\n4. Update XML libraries to latest versions\n5. For Java: XMLConstants.FEATURE_SECURE_PROCESSING = true\n6. For Python: Use defusedxml library\n7. Implement input validation\n8. Use XML schema validation'
            })
        else:
            results.append({
                'Test': 'XML External Entity - XXE (OWASP A05:2021)',
                'Finding': 'No XXE vulnerability detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'XML External Entity - XXE (OWASP A05:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_command_injection(session, url, path):
    """Test Command/Script Injection"""
    results = []
    
    try:
        command_payloads = [
            '; ls -la',
            '| whoami',
            '& dir',
            '`id`',
            '$(whoami)',
            '; cat /etc/passwd'
        ]
        
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            
            for param_name in params.keys():
                for payload in command_payloads[:2]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, verify=False)
                        # Check for command execution indicators
                        indicators = ['root:', 'bin/bash', 'uid=', 'total', 'drwxr']
                        if any(indicator in response.text for indicator in indicators):
                            vulnerable = True
                            break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'Command/Script Injection',
                'Finding': 'Potential command injection vulnerability detected - server may execute user input as system commands',
                'Severity': 'Critical',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Never execute user input as system commands. Use parameterized APIs instead.',
                'Resolution Steps': '1. Never pass user input to system shells\n2. Use safe APIs (subprocess with shell=False)\n3. Implement strict input validation with allowlist\n4. Use least privilege principles\n5. Disable dangerous functions (exec, eval, system)\n6. Implement command allowlisting\n7. Use containerization/sandboxing\n8. Regular code review and security testing'
            })
        else:
            results.append({
                'Test': 'Command/Script Injection',
                'Finding': 'No command injection vulnerability detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Command/Script Injection',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_hardcoded_secrets(session, url, path):
    """Test for hard-coded secrets in response"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        
        # Patterns to detect secrets
        secret_patterns = {
            'api_key': r'api[_-]?key[\s:=]+["\']?([a-zA-Z0-9_\-]{20,})',
            'secret': r'secret[\s:=]+["\']?([a-zA-Z0-9_\-]{20,})',
            'password': r'password[\s:=]+["\']([^"\']{6,})',
            'token': r'token[\s:=]+["\']?([a-zA-Z0-9_\-]{20,})',
            'aws_key': r'AKIA[0-9A-Z]{16}',
            'private_key': r'-----BEGIN (RSA|DSA|EC) PRIVATE KEY-----'
        }
        
        found_secrets = []
        for secret_type, pattern in secret_patterns.items():
            matches = re.findall(pattern, response.text, re.IGNORECASE)
            if matches:
                found_secrets.append(secret_type)
        
        if found_secrets:
            results.append({
                'Test': 'Hard-Coded Secrets Detection',
                'Finding': f'Potential hard-coded secrets found in response: {", ".join(found_secrets)}',
                'Severity': 'Critical',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Remove all hard-coded credentials and secrets from code and responses.',
                'Resolution Steps': '1. Remove all hard-coded credentials from code\n2. Use environment variables for secrets\n3. Implement secret management systems (Vault, AWS Secrets Manager)\n4. Never commit secrets to version control\n5. Use .gitignore for sensitive files\n6. Scan code with secret detection tools\n7. Rotate exposed credentials immediately\n8. Implement proper access controls'
            })
        else:
            results.append({
                'Test': 'Hard-Coded Secrets Detection',
                'Finding': 'No obvious hard-coded secrets detected in response',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Hard-Coded Secrets Detection',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_session_timeout(session, url, path):
    """Test session timeout implementation"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        
        # Check for session timeout mechanisms
        has_timeout = False
        timeout_indicators = []
        
        # Check cookies for max-age or expires
        cookies = session.cookies
        for cookie in cookies:
            if cookie.expires:
                has_timeout = True
                timeout_indicators.append(f"Cookie '{cookie.name}' has expiration")
        
        # Check for session management headers
        cache_control = response.headers.get('Cache-Control', '')
        if 'no-cache' in cache_control or 'private' in cache_control:
            timeout_indicators.append('Cache-Control headers present')
        
        if not has_timeout and not timeout_indicators:
            results.append({
                'Test': 'Session Timeout Testing',
                'Finding': 'No session timeout mechanism detected - sessions may persist indefinitely',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement proper session timeout and expiration.',
                'Resolution Steps': '1. Implement idle timeout (15-30 minutes)\n2. Implement absolute timeout (2-8 hours)\n3. Set cookie expiration times\n4. Clear sessions server-side on logout\n5. Implement session invalidation\n6. Use secure session management libraries\n7. Warn users before timeout\n8. Re-authenticate for sensitive operations'
            })
        else:
            results.append({
                'Test': 'Session Timeout Testing',
                'Finding': f'Session management appears configured: {", ".join(timeout_indicators) if timeout_indicators else "timeout detected"}',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Session Timeout Testing',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_open_redirect(session, url, path):
    """Test for open redirection vulnerabilities"""
    results = []
    
    try:
        redirect_payloads = [
            'http://evil.com',
            'https://malicious.example.com',
            '//evil.com',
            'javascript:alert(1)'
        ]
        
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            redirect_params = [p for p in params if any(k in p.lower() for k in ['redirect', 'url', 'next', 'return', 'dest', 'redir'])]
            
            for param_name in redirect_params:
                for payload in redirect_payloads[:2]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, verify=False, allow_redirects=False)
                        location = response.headers.get('Location', '')
                        if payload in location or 'evil.com' in location:
                            vulnerable = True
                            break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'Open Redirection',
                'Finding': 'Open redirect vulnerability detected - application redirects to user-supplied URLs',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Validate all redirect URLs against an allowlist of permitted destinations.',
                'Resolution Steps': '1. Implement URL allowlist for redirects\n2. Validate redirect URLs server-side\n3. Use relative paths for internal redirects\n4. Avoid user-controlled redirect parameters\n5. Display warning for external redirects\n6. Implement CSRF tokens for redirects\n7. Log all redirect attempts\n8. Use indirect references (mapping IDs)'
            })
        else:
            results.append({
                'Test': 'Open Redirection',
                'Finding': 'No open redirect vulnerability detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Open Redirection',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_outdated_components(session, url, path):
    """Test for outdated components and libraries"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        
        # Check server header for version
        server_header = response.headers.get('Server', '')
        x_powered = response.headers.get('X-Powered-By', '')
        
        outdated_indicators = []
        
        # Check for version numbers in headers
        version_patterns = r'(\d+\.\d+\.\d+)'
        
        if server_header:
            versions = re.findall(version_patterns, server_header)
            if versions:
                outdated_indicators.append(f'Server version disclosed: {server_header}')
        
        if x_powered:
            outdated_indicators.append(f'X-Powered-By header present: {x_powered}')
        
        # Check response for common library indicators
        library_patterns = {
            'jQuery': r'jquery[/-](\d+\.\d+\.\d+)',
            'Bootstrap': r'bootstrap[/-](\d+\.\d+\.\d+)',
            'Angular': r'angular[/-](\d+\.\d+\.\d+)',
            'React': r'react[/-](\d+\.\d+\.\d+)'
        }
        
        for lib, pattern in library_patterns.items():
            matches = re.findall(pattern, response.text, re.IGNORECASE)
            if matches:
                outdated_indicators.append(f'{lib} version {matches[0]} detected')
        
        if outdated_indicators:
            results.append({
                'Test': 'Outdated Components Detection',
                'Finding': f'Version information exposed: {", ".join(outdated_indicators[:3])}. May contain known vulnerabilities.',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Update all components to latest versions and remove version disclosure.',
                'Resolution Steps': '1. Inventory all components and dependencies\n2. Update to latest stable versions\n3. Remove version headers (Server, X-Powered-By)\n4. Subscribe to security advisories\n5. Implement automated dependency scanning\n6. Use OWASP Dependency-Check\n7. Regular update schedule\n8. Test updates in staging first'
            })
        else:
            results.append({
                'Test': 'Outdated Components Detection',
                'Finding': 'No obvious outdated component indicators detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Outdated Components Detection',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_verbose_errors(session, url, path):
    """Test for verbose error messages"""
    results = []
    
    try:
        # Test with various invalid inputs to trigger errors
        error_triggers = [
            ('invalid_param', '"><script>alert(1)</script>'),
            ('id', '-1'),
            ('file', '../../../etc/passwd'),
            ('search', "' OR '1'='1")
        ]
        
        verbose_found = False
        error_details = []
        
        for param, value in error_triggers:
            try:
                test_url = f"{url}?{param}={value}"
                response = session.get(test_url, timeout=10, verify=False)
                
                # Check for verbose error indicators
                error_patterns = [
                    r'(mysql|postgres|oracle|sql).*error',
                    r'syntax error',
                    r'stack trace',
                    r'exception',
                    r'warning:',
                    r'fatal error',
                    r'<b>.*error.*</b>',
                    r'on line \d+',
                    r'in /.*\.php'
                ]
                
                for pattern in error_patterns:
                    if re.search(pattern, response.text, re.IGNORECASE):
                        verbose_found = True
                        error_details.append(pattern.replace('\\', ''))
                        break
                        
            except:
                continue
        
        if verbose_found:
            results.append({
                'Test': 'Verbose Error Messages',
                'Finding': f'Verbose error messages detected in server responses. Details: {", ".join(set(error_details)[:2])}',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement custom error pages and disable detailed error messages in production.',
                'Resolution Steps': '1. Disable debug mode in production\n2. Implement custom error pages\n3. Log detailed errors server-side only\n4. Return generic error messages to users\n5. Configure web server error pages\n6. Remove stack traces from responses\n7. Sanitize error messages\n8. Implement centralized error handling'
            })
        else:
            results.append({
                'Test': 'Verbose Error Messages',
                'Finding': 'No verbose error messages detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Verbose Error Messages',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_rate_limiting(session, url, path):
    """Test for rate limiting and throttling"""
    results = []
    
    try:
        # Make multiple rapid requests
        rapid_requests = 10
        responses = []
        
        for i in range(rapid_requests):
            try:
                response = session.get(url, timeout=5, verify=False)
                responses.append(response.status_code)
            except:
                break
        
        # Check if rate limiting is in place
        rate_limited = any(status == 429 for status in responses)
        has_retry_after = False
        
        if responses:
            last_response = session.get(url, timeout=10, verify=False)
            has_retry_after = 'Retry-After' in last_response.headers or 'X-RateLimit-Limit' in last_response.headers
        
        if not rate_limited and not has_retry_after:
            results.append({
                'Test': 'Rate Limiting & Throttling',
                'Finding': 'No rate limiting detected - application accepts unlimited rapid requests',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement rate limiting to prevent abuse and DoS attacks.',
                'Resolution Steps': '1. Implement request rate limiting per IP/user\n2. Use token bucket or leaky bucket algorithm\n3. Set appropriate limits (e.g., 100 req/min)\n4. Return 429 Too Many Requests status\n5. Include Retry-After header\n6. Implement CAPTCHA for repeated violations\n7. Use WAF with rate limiting\n8. Monitor and log excessive requests'
            })
        else:
            status_msg = 'Rate limiting detected (429 status)' if rate_limited else 'Rate limiting headers present'
            results.append({
                'Test': 'Rate Limiting & Throttling',
                'Finding': f'{status_msg}',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Rate Limiting & Throttling',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_csp_header(session, url, path):
    """Test Content Security Policy (CSP) header"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        
        csp_header = response.headers.get('Content-Security-Policy', '')
        csp_report = response.headers.get('Content-Security-Policy-Report-Only', '')
        
        if not csp_header and not csp_report:
            results.append({
                'Test': 'Content Security Policy (CSP)',
                'Finding': 'No Content-Security-Policy header detected - missing defense against XSS and injection attacks',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement Content-Security-Policy header to prevent XSS and data injection.',
                'Resolution Steps': "1. Implement CSP header with restrictive policy\n2. Use 'self' for script sources\n3. Avoid 'unsafe-inline' and 'unsafe-eval'\n4. Use nonces or hashes for inline scripts\n5. Enable CSP reporting\n6. Test with CSP-Report-Only first\n7. Regularly review and update policy\n8. Use CSP Level 3 features where supported"
            })
        else:
            # Check for weak CSP
            weak_indicators = []
            full_csp = csp_header or csp_report
            
            if 'unsafe-inline' in full_csp:
                weak_indicators.append("'unsafe-inline' allows inline scripts")
            if 'unsafe-eval' in full_csp:
                weak_indicators.append("'unsafe-eval' allows eval()")
            if '*' in full_csp:
                weak_indicators.append("wildcard (*) allows any source")
            
            if weak_indicators:
                results.append({
                    'Test': 'Content Security Policy (CSP)',
                    'Finding': f'CSP header present but potentially weak: {", ".join(weak_indicators)}',
                    'Severity': 'Low',
                    'Status': 'Vulnerable',
                    'Vulnerable Path': url,
                    'Remediation': 'Strengthen CSP policy by removing unsafe directives.',
                    'Resolution Steps': "1. Remove 'unsafe-inline' and 'unsafe-eval'\n2. Replace wildcards with specific domains\n3. Use nonces for inline scripts\n4. Implement strict-dynamic\n5. Enable upgrade-insecure-requests\n6. Test policy thoroughly\n7. Monitor CSP reports\n8. Use CSP validator tools"
                })
            else:
                results.append({
                    'Test': 'Content Security Policy (CSP)',
                    'Finding': 'Content-Security-Policy header is properly configured',
                    'Severity': 'Info',
                    'Status': 'Secure',
                    'Vulnerable Path': 'N/A',
                    'Remediation': 'N/A',
                    'Resolution Steps': 'N/A'
                })
    
    except Exception as e:
        results.append({
            'Test': 'Content Security Policy (CSP)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_invalid_characters(session, url, path):
    """Test if invalid characters are allowed in form fields"""
    results = []
    
    try:
        # Get the page and look for forms
        response = session.get(url, timeout=10, verify=False)
        soup = BeautifulSoup(response.text, 'html.parser')
        forms = soup.find_all('form')
        
        if not forms:
            results.append({
                'Test': 'Invalid Characters in Form Fields',
                'Finding': 'No forms detected on this page',
                'Severity': 'Info',
                'Status': 'Complete',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
            return results
        
        # Test with invalid characters
        invalid_chars = ['<script>', '; DROP TABLE', '../../../', 'null\x00byte', '${7*7}']
        
        vulnerable = False
        for form in forms[:1]:  # Test first form only
            inputs = form.find_all('input')
            if not inputs:
                continue
            
            # Build form data
            form_data = {}
            for inp in inputs:
                name = inp.get('name')
                if name and inp.get('type') != 'submit':
                    form_data[name] = invalid_chars[0]
            
            if form_data:
                action = form.get('action', url)
                method = form.get('method', 'get').lower()
                
                try:
                    if method == 'post':
                        test_response = session.post(urljoin(url, action), data=form_data, timeout=10, verify=False)
                    else:
                        test_response = session.get(urljoin(url, action), params=form_data, timeout=10, verify=False)
                    
                    # Check if invalid input is reflected or accepted
                    if invalid_chars[0] in test_response.text:
                        vulnerable = True
                except:
                    pass
        
        if vulnerable:
            results.append({
                'Test': 'Invalid Characters in Form Fields',
                'Finding': 'Form accepts and reflects invalid/dangerous characters without proper validation',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement strict input validation with allowlist approach.',
                'Resolution Steps': '1. Implement input validation for all fields\n2. Use allowlist (whitelist) approach\n3. Reject dangerous characters\n4. Sanitize input before processing\n5. Use framework validation libraries\n6. Implement server-side validation\n7. Set maximum length limits\n8. Encode output properly'
            })
        else:
            results.append({
                'Test': 'Invalid Characters in Form Fields',
                'Finding': 'Form validation appears to be implemented',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Invalid Characters in Form Fields',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_waf_detection(session, url, path):
    """Test for WAF (Web Application Firewall) presence"""
    results = []
    
    try:
        # Send malicious payloads to trigger WAF
        waf_payloads = [
            "' OR '1'='1",
            "<script>alert('XSS')</script>",
            "../../../../etc/passwd",
            "<?php system('id'); ?>"
        ]
        
        waf_detected = False
        waf_indicators = []
        
        for payload in waf_payloads:
            try:
                test_url = f"{url}?test={payload}"
                response = session.get(test_url, timeout=10, verify=False)
                
                # Check for WAF response codes
                if response.status_code in [403, 406, 419, 429, 501, 503]:
                    waf_detected = True
                    waf_indicators.append(f'Status {response.status_code}')
                
                # Check for WAF headers
                waf_headers = {
                    'cloudflare': ['cf-ray', 'cf-request-id'],
                    'akamai': ['akamai-ghost'],
                    'aws-waf': ['x-amzn-requestid', 'x-amz-cf-id'],
                    'imperva': ['x-iinfo'],
                    'barracuda': ['barra_counter_session'],
                    'f5': ['x-cnection'],
                    'sucuri': ['x-sucuri-id']
                }
                
                for waf_name, headers in waf_headers.items():
                    if any(h.lower() in [k.lower() for k in response.headers.keys()] for h in headers):
                        waf_detected = True
                        waf_indicators.append(f'{waf_name.upper()} detected')
                
                # Check response body for WAF messages
                waf_signatures = [
                    'Access Denied',
                    'Blocked by',
                    'Web Application Firewall',
                    'Security Policy',
                    'Request Rejected',
                    'Forbidden'
                ]
                
                for signature in waf_signatures:
                    if signature.lower() in response.text.lower():
                        waf_detected = True
                        waf_indicators.append(f'WAF message: {signature}')
                        break
                
                if waf_detected:
                    break
                    
            except:
                continue
        
        if waf_detected:
            results.append({
                'Test': 'WAF Detection & Awareness',
                'Finding': f'Web Application Firewall detected: {", ".join(set(waf_indicators)[:2])}. Additional testing may be blocked.',
                'Severity': 'Info',
                'Status': 'Complete',
                'Vulnerable Path': url,
                'Remediation': 'WAF is a security control. Ensure it is properly configured.',
                'Resolution Steps': '1. Verify WAF rules are up to date\n2. Test WAF in audit mode first\n3. Configure custom rules for your app\n4. Monitor WAF logs regularly\n5. Fine-tune to reduce false positives\n6. Keep WAF signatures updated\n7. Use geo-blocking where appropriate\n8. Implement rate limiting rules'
            })
        else:
            results.append({
                'Test': 'WAF Detection & Awareness',
                'Finding': 'No Web Application Firewall detected - consider implementing one for additional security',
                'Severity': 'Low',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Consider implementing a Web Application Firewall.',
                'Resolution Steps': '1. Evaluate WAF solutions (CloudFlare, AWS WAF, Imperva)\n2. Implement WAF in monitoring mode first\n3. Configure OWASP Core Rule Set\n4. Create custom rules for your application\n5. Enable request logging\n6. Set up alerts for attacks\n7. Regularly review and update rules\n8. Train team on WAF management'
            })
    
    except Exception as e:
        results.append({
            'Test': 'WAF Detection & Awareness',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_session_timeout_owasp(session, url, path):
    """Test Session Timeout (OWASP A07:2021)"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        
        session_cookies = [c for c in session.cookies if any(n in c.name.lower() for n in ['session', 'sess', 'jsessionid', 'phpsessid'])]
        
        if session_cookies:
            set_cookie = response.headers.get('Set-Cookie', '')
            has_timeout = 'max-age=' in set_cookie.lower() or 'expires=' in set_cookie.lower()
            
            if not has_timeout:
                results.append({
                    'Test': 'Session Timeout (OWASP A07:2021)',
                    'Finding': 'Session cookies without explicit timeout/expiration',
                    'Severity': 'Medium',
                    'Status': 'Vulnerable',
                    'Vulnerable Path': url,
                    'Remediation': 'Implement proper session timeout with idle and absolute timeouts.',
                    'Resolution Steps': '1. Set idle timeout: 15-30 minutes of inactivity\n2. Set absolute timeout: 2-8 hours maximum\n3. Implement sliding expiration for active sessions\n4. Clear session data on timeout\n5. Redirect to login on timeout\n6. Use secure server-side session storage\n7. Implement separate "Remember Me" functionality\n8. Log session timeout events for monitoring'
                })
            else:
                results.append({
                    'Test': 'Session Timeout (OWASP A07:2021)',
                    'Finding': 'Session timeout is configured',
                    'Severity': 'Info',
                    'Status': 'Secure',
                    'Vulnerable Path': 'N/A',
                    'Remediation': 'N/A',
                    'Resolution Steps': 'N/A'
                })
        else:
            results.append({
                'Test': 'Session Timeout (OWASP A07:2021)',
                'Finding': 'No session cookies detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Session Timeout (OWASP A07:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_open_redirect_owasp(session, url, path):
    """Test Open Redirection (OWASP A01:2021)"""
    results = []
    
    try:
        redirect_payloads = ['http://evil.com', '//evil.com']
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            redirect_params = [p for p in params if any(k in p.lower() for k in ['redirect', 'url', 'return', 'next', 'goto'])]
            
            for param_name in redirect_params:
                for payload in redirect_payloads[:1]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, allow_redirects=False, verify=False)
                        if response.status_code in [301, 302, 303, 307, 308]:
                            location = response.headers.get('Location', '')
                            if 'evil.com' in location:
                                vulnerable = True
                                break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'Open Redirection (OWASP A01:2021)',
                'Finding': 'Open redirect vulnerability - application redirects to arbitrary URLs',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Validate redirect URLs against allowlist of permitted domains.',
                'Resolution Steps': '1. Implement allowlist of permitted redirect domains\n2. Validate all redirect URLs server-side\n3. Use relative URLs for internal redirects\n4. Avoid user input in redirect parameters\n5. Use indirect reference maps\n6. Display warning for external redirects\n7. Implement CSRF protection for redirects\n8. Log all redirect attempts'
            })
        else:
            results.append({
                'Test': 'Open Redirection (OWASP A01:2021)',
                'Finding': 'No open redirect vulnerability detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Open Redirection (OWASP A01:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_outdated_components_owasp(session, url, path):
    """Test Outdated Components (OWASP A06:2021)"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        headers = response.headers
        content = response.text
        
        outdated = []
        
        # Check headers
        server = headers.get('Server', '')
        powered_by = headers.get('X-Powered-By', '')
        
        version_checks = [
            (r'Apache/2\.(0|2)', 'Apache 2.0/2.2 (EOL)'),
            (r'PHP/5\.', 'PHP 5.x (EOL)'),
            (r'PHP/7\.[0-2]', 'PHP 7.0-7.2 (EOL)'),
        ]
        
        for pattern, desc in version_checks:
            if re.search(pattern, server + powered_by):
                outdated.append(desc)
        
        # Check content for library versions
        lib_checks = [
            (r'jquery[/-]1\.[0-8]', 'jQuery 1.x (outdated)'),
            (r'angular\.js[/-]1\.[0-4]', 'AngularJS 1.x (outdated)'),
        ]
        
        for pattern, desc in lib_checks:
            if re.search(pattern, content, re.I):
                outdated.append(desc)
        
        if outdated:
            results.append({
                'Test': 'Outdated Components (OWASP A06:2021)',
                'Finding': f'Outdated components detected: {", ".join(outdated)}',
                'Severity': 'High',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Update all components to latest stable versions.',
                'Resolution Steps': '1. Inventory all components and versions\n2. Check for security advisories\n3. Update to latest stable versions\n4. Update dependencies (npm audit, pip-audit)\n5. Automate dependency scanning\n6. Subscribe to security notifications\n7. Use tools: Snyk, Dependabot, OWASP Dependency-Check\n8. Establish regular update schedule'
            })
        else:
            results.append({
                'Test': 'Outdated Components (OWASP A06:2021)',
                'Finding': 'No obviously outdated components detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Outdated Components (OWASP A06:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_sensitive_data_owasp(session, url, path):
    """Test Sensitive Data Exposure (OWASP A02:2021)"""
    results = []
    
    try:
        response = session.get(url, timeout=10, verify=False)
        content = response.text.lower()
        
        sensitive_patterns = [
            (r'\b\d{3}-\d{2}-\d{4}\b', 'SSN'),
            (r'\b\d{4}[- ]?\d{4}[- ]?\d{4}[- ]?\d{4}\b', 'Credit Card'),
            (r'password\s*[:=]\s*["\']?[^\s"\']+', 'Password'),
        ]
        
        found = []
        
        for pattern, dtype in sensitive_patterns:
            if re.search(pattern, content):
                found.append(dtype)
        
        is_https = url.startswith('https://')
        
        if found or not is_https:
            severity = 'Critical' if found else 'High'
            finding_parts = []
            
            if found:
                finding_parts.append(f'Sensitive data in response: {", ".join(found)}')
            if not is_https:
                finding_parts.append('Unencrypted HTTP connection')
            
            results.append({
                'Test': 'Sensitive Data Exposure (OWASP A02:2021)',
                'Finding': '. '.join(finding_parts),
                'Severity': severity,
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Encrypt data in transit and at rest. Remove sensitive data from responses.',
                'Resolution Steps': '1. Enable HTTPS for entire site\n2. Remove sensitive data from HTML/JavaScript\n3. Implement data masking (show last 4 digits only)\n4. Use POST for sensitive data (never GET)\n5. Implement field-level encryption\n6. Set Cache-Control: no-store for sensitive pages\n7. Use secure session management\n8. Regular data handling audits'
            })
        else:
            results.append({
                'Test': 'Sensitive Data Exposure (OWASP A02:2021)',
                'Finding': 'No obvious sensitive data exposure detected',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Sensitive Data Exposure (OWASP A02:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_invalid_input_owasp(session, url, path):
    """Test Invalid Input Handling (OWASP A03:2021)"""
    results = []
    
    try:
        invalid_payloads = ['<script>', '../../', '${7*7}', '\x00\x01']
        vulnerable = False
        
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query)
            
            for param_name in params.keys():
                for payload in invalid_payloads[:1]:
                    test_params = params.copy()
                    test_params[param_name] = [payload]
                    test_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(test_params, doseq=True)}"
                    
                    try:
                        response = session.get(test_url, timeout=10, verify=False)
                        if response.status_code != 400 and payload in response.text:
                            vulnerable = True
                            break
                    except:
                        continue
                if vulnerable:
                    break
        
        if vulnerable:
            results.append({
                'Test': 'Invalid Input Handling (OWASP A03:2021)',
                'Finding': 'Application accepts invalid/dangerous characters without proper validation',
                'Severity': 'Medium',
                'Status': 'Vulnerable',
                'Vulnerable Path': url,
                'Remediation': 'Implement strict input validation with allowlisting.',
                'Resolution Steps': '1. Define allowed character sets per field\n2. Implement server-side validation (whitelist approach)\n3. Reject invalid input with 400 status\n4. Sanitize before processing\n5. Use validation libraries (Joi, express-validator)\n6. Implement length restrictions\n7. Encode special characters in output\n8. Log validation failures'
            })
        else:
            results.append({
                'Test': 'Invalid Input Handling (OWASP A03:2021)',
                'Finding': 'Input validation appears to be implemented',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'Invalid Input Handling (OWASP A03:2021)',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

def test_waf_detection(session, url, path):
    """Test WAF Detection"""
    results = []
    
    try:
        malicious_payload = "' OR '1'='1 UNION SELECT NULL--"
        test_url = f"{url}?test={malicious_payload}"
        
        try:
            response = session.get(test_url, timeout=10, verify=False)
            
            waf_signatures = {
                'Cloudflare': ['cloudflare', 'cf-ray', '__cfduid'],
                'AWS WAF': ['x-amzn', 'x-amz-'],
                'Akamai': ['akamai'],
                'Imperva': ['incap_ses', 'visid_incap'],
            }
            
            detected_waf = None
            
            for waf_name, signatures in waf_signatures.items():
                for sig in signatures:
                    if any(sig in h.lower() for h in response.headers.keys()):
                        detected_waf = waf_name
                        break
                if detected_waf:
                    break
            
            if response.status_code in [403, 406, 419, 429]:
                if not detected_waf:
                    detected_waf = f"WAF (Status {response.status_code})"
            
            if detected_waf:
                results.append({
                    'Test': 'WAF Detection',
                    'Finding': f'Web Application Firewall detected: {detected_waf}',
                    'Severity': 'Info',
                    'Status': 'Secure',
                    'Vulnerable Path': 'N/A',
                    'Remediation': 'N/A',
                    'Resolution Steps': 'N/A'
                })
            else:
                results.append({
                    'Test': 'WAF Detection',
                    'Finding': 'No WAF detected - consider implementing for additional protection',
                    'Severity': 'Low',
                    'Status': 'Vulnerable',
                    'Vulnerable Path': url,
                    'Remediation': 'Implement Web Application Firewall for protection against attacks.',
                    'Resolution Steps': '1. Evaluate WAF solutions (Cloudflare, AWS WAF, ModSecurity)\n2. Deploy cloud-based or on-premise WAF\n3. Configure OWASP Core Rule Set\n4. Enable protection for OWASP Top 10\n5. Customize rules for application\n6. Enable logging and monitoring\n7. Set up attack alerts\n8. Regular rule updates and testing'
                })
        
        except:
            results.append({
                'Test': 'WAF Detection',
                'Finding': 'Request blocked - possible WAF present',
                'Severity': 'Info',
                'Status': 'Secure',
                'Vulnerable Path': 'N/A',
                'Remediation': 'N/A',
                'Resolution Steps': 'N/A'
            })
    
    except Exception as e:
        results.append({
            'Test': 'WAF Detection',
            'Finding': f'Error testing: {str(e)}',
            'Severity': 'Info',
            'Status': 'Error',
            'Vulnerable Path': 'N/A',
            'Remediation': 'N/A',
            'Resolution Steps': 'N/A'
        })
    
    return results

# ==================== PATH GROUPING AND SELECTION ====================

def group_and_select_paths(paths):
    """
    Intelligently group paths to avoid testing too many similar IDs
    - Tests ALL unique path patterns (campaigns, videos, dashboard, users, etc.)
    - For paths with IDs, tests only 2 representative IDs max
    - Each selected path will be tested and shown separately in report
    """
    from collections import defaultdict
    import re
    
    print("\n[+] Applying intelligent path selection...")
    
    # Dictionary to store path groups by pattern
    path_groups = defaultdict(list)
    
    # Group paths by their pattern
    for path in paths:
        # Remove query parameters for grouping
        clean_path = path.split('?')[0]
        
        # Create pattern by replacing numeric IDs with {id}
        # This groups /videos/1, /videos/2, /videos/3 together
        pattern = re.sub(r'/\d+(/|$)', '/{id}/', clean_path)
        pattern = re.sub(r'/\d+$', '/{id}', pattern)
        
        # Group by this pattern
        path_groups[pattern].append(path)
    
    # Select paths from each group
    selected_paths = []
    
    print("\n[+] Path Selection Summary:")
    for pattern, group_paths in sorted(path_groups.items()):
        group_paths.sort()
        
        if len(group_paths) == 1:
            # Only one path, select it
            selected_paths.append(group_paths[0])
            print(f"  ✓ {pattern}: 1 path selected")
        
        elif len(group_paths) == 2:
            # Two paths, select both
            selected_paths.extend(group_paths[:2])
            print(f"  ✓ {pattern}: 2 paths selected")
        
        else:
            # Multiple paths with IDs - select first 2 as representatives
            selected_paths.extend(group_paths[:2])
            print(f"  ✓ {pattern}: 2 of {len(group_paths)} paths selected (skipped {len(group_paths)-2} similar IDs)")
    
    print(f"\n[+] Total: Selected {len(selected_paths)} paths from {len(paths)} discovered paths")
    print(f"[+] Each selected path will be tested individually and shown in the report\n")
    
    return selected_paths

# ==================== COMPREHENSIVE PATH TESTING ====================

def test_path_owasp_complete(base_url, paths, auth_credentials=None):
    """
    Run all OWASP tests on discovered paths
    Each path is tested individually and results are shown separately in report
    No grouping - every path gets its own vulnerability entries
    """
    print(f"\n[+] Running comprehensive OWASP tests on discovered paths...")
    
    # Apply intelligent path selection to limit ID testing
    selected_paths = group_and_select_paths(paths)
    
    # List to collect all results - each path tested separately
    all_results = []
    
    session = requests.Session()
    session.verify = False
    
    if auth_credentials:
        auth_type = auth_credentials.get('type')
        auth_data = auth_credentials.get('data', {})
        
        if auth_type == 'basic':
            session.auth = (auth_data.get('username'), auth_data.get('password'))
    
    # Define all test functions
    test_functions = [
        ('Rate Limiting & Throttling', test_rate_limiting),
        ('Content Security Policy (CSP)', test_csp_header),
        ('Hard-Coded Secrets', test_hardcoded_secrets),
        ('SQL Injection', test_sql_injection_owasp),
        ('Command/Script Injection', test_command_injection),
        ('Cross-Site Scripting (XSS)', test_xss_owasp),
        ('CSRF Protection', test_csrf_owasp),
        ('SSRF', test_ssrf_owasp),
        ('XXE Injection', test_xxe_owasp),
        ('Session Timeout', test_session_timeout),
        ('Open Redirect', test_open_redirect),
        ('Outdated Components', test_outdated_components),
        ('Verbose Error Messages', test_verbose_errors),
        ('Invalid Characters in Forms', test_invalid_characters),
        ('WAF Detection', test_waf_detection)
    ]
    
    # Test each selected path individually
    for i, path in enumerate(selected_paths, 1):
        url = path if path.startswith('http') else urljoin(base_url, path)
        print(f"\n  [{i}/{len(selected_paths)}] Testing: {url}")
        
        # Run all tests on this path
        for test_name, test_func in test_functions:
            try:
                # Get results for this specific path
                path_results = test_func(session, url, path)
                
                # Add all results directly - no grouping
                # Each path will have its own set of test results
                all_results.extend(path_results)
                            
            except Exception as e:
                # Silently continue on errors to avoid cluttering output
                continue
    
    print(f"\n[+] Completed OWASP testing")
    print(f"[+] Tested {len(selected_paths)} unique paths")
    print(f"[+] Generated {len(all_results)} total test results")
    
    return all_results
# ==================== MAIN SCAN FUNCTION ====================

def perform_vapt_scan(target, auth_credentials=None, owasp_enabled=True, progress_callback=None):
    """Main VAPT scan function with real-time progress updates"""
    try:
        print(f"\n{'='*70}")
        print(f"COMPREHENSIVE WEB APPLICATION PENETRATION TESTING")
        print(f"Target: {target}")
        print(f"{'='*70}\n")
        
        all_results = []
        discovered_paths = []
        
        # Phase 1: Network Testing
        print(f"\n{'='*70}")
        print(f"PHASE 1: NETWORK SECURITY TESTING")
        print(f"{'='*70}")
        
        if progress_callback:
            progress_callback({'type': 'phase', 'phase': 1, 'name': 'Network Security Testing'})
        
        all_results.extend(test_reconnaissance(target))
        all_results.extend(test_port_scanning(target))
        all_results.extend(test_service_detection(target))
        all_results.extend(test_vulnerability_scanning(target))
        
        # Phase 2: Web Application Testing
        is_web_app = check_web_application(target)
        
        if is_web_app:
            print(f"\n{'='*70}")
            print(f"PHASE 2: WEB CRAWLING & PATH DISCOVERY")
            print(f"{'='*70}")
            
            if progress_callback:
                progress_callback({'type': 'phase', 'phase': 2, 'name': 'Web Crawling & Path Discovery'})
            
            base_url = get_base_url(target)
            crawl_data = crawl_website(base_url, auth_credentials, max_pages=50, progress_callback=progress_callback)
            discovered_paths = crawl_data['paths']
            
            print(f"\n[+] Discovered {len(discovered_paths)} unique paths")
            
            # Phase 3: OWASP Testing on Each Path
            if owasp_enabled and discovered_paths:
                print(f"\n{'='*70}")
                print(f"PHASE 3: OWASP TOP 10 VULNERABILITY TESTING")
                print(f"{'='*70}")
                
                if progress_callback:
                    progress_callback({'type': 'phase', 'phase': 3, 'name': 'OWASP Top 10 Vulnerability Testing'})
                
                owasp_results = test_path_owasp_complete(base_url, discovered_paths, auth_credentials)
                all_results.extend(owasp_results)
        else:
            print("\n[!] Target is not a web application - skipping web tests")
        
        # Generate Excel Report
        if progress_callback:
            progress_callback({'type': 'phase', 'phase': 4, 'name': 'Generating Report'})
        
        filename = generate_excel_report(target, all_results, discovered_paths)
        
        print(f"\n{'='*70}")
        print(f"SCAN COMPLETE!")
        print(f"Report: {filename}")
        print(f"{'='*70}\n")
        
        return {
            'status': 'success',
            'filename': filename,
            'results': all_results
        }
        
    except Exception as e:
        print(f"\n[!] Scan error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'status': 'error',
            'message': str(e)
        }

def generate_excel_report(target, results, discovered_paths):
    """Generate Excel report matching the exact format"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    clean_target = target.replace('/', '_').replace(':', '_').replace('?', '_')
    filename = f"VAPT_Report_{clean_target}_{timestamp}.xlsx"
    
    # Create DataFrame
    df = pd.DataFrame(results)
    
    # Reorder columns to match exact format
    column_order = ['Test', 'Severity', 'Status', 'Finding', 'Vulnerable Path', 'Remediation', 'Resolution Steps']
    df = df[column_order]
    
    # Create Excel with two sheets
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Scan Results', index=False)
        
        # Create Discovered Paths sheet
        if discovered_paths:
            paths_df = pd.DataFrame({
                'Discovered Paths': discovered_paths,
                'Total Paths': [len(discovered_paths)] * len(discovered_paths)
            })
            paths_df.to_excel(writer, sheet_name='Discovered Paths', index=False)
    
    # Format the workbook
    wb = load_workbook(filename)
    ws = wb['Scan Results']
    
    # Define colors
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    medium_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    low_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    info_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Define fonts
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Format data rows
    for row_idx in range(2, ws.max_row + 1):
        severity = ws[f'B{row_idx}'].value
        
        # Apply severity colors
        if severity == 'Critical':
            ws[f'B{row_idx}'].fill = critical_fill
            ws[f'B{row_idx}'].font = Font(bold=True, color="FFFFFF")
        elif severity == 'High':
            ws[f'B{row_idx}'].fill = high_fill
            ws[f'B{row_idx}'].font = Font(bold=True)
        elif severity == 'Medium':
            ws[f'B{row_idx}'].fill = medium_fill
            ws[f'B{row_idx}'].font = Font(bold=True)
        elif severity == 'Low':
            ws[f'B{row_idx}'].fill = low_fill
            ws[f'B{row_idx}'].font = Font(bold=True)
        else:
            ws[f'B{row_idx}'].fill = info_fill
            ws[f'B{row_idx}'].font = Font(bold=True)
        
        # Apply borders and alignment to all cells
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 40  # Test
    ws.column_dimensions['B'].width = 12  # Severity
    ws.column_dimensions['C'].width = 12  # Status
    ws.column_dimensions['D'].width = 60  # Finding
    ws.column_dimensions['E'].width = 50  # Vulnerable Path
    ws.column_dimensions['F'].width = 70  # Remediation
    ws.column_dimensions['G'].width = 80  # Resolution Steps
    
    # Set row heights
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 100
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    
    # Format Discovered Paths sheet if it exists
    if 'Discovered Paths' in wb.sheetnames:
        paths_ws = wb['Discovered Paths']
        
        # Format header
        for cell in paths_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Format data
        for row_idx in range(2, paths_ws.max_row + 1):
            for col in range(1, 3):
                cell = paths_ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        paths_ws.column_dimensions['A'].width = 80
        paths_ws.column_dimensions['B'].width = 15
        paths_ws.freeze_panes = 'A2'
    
    wb.save(filename)
    print(f"[+] Report saved: {filename}")
    
    return filename

# ==================== ENTRY POINT ====================

if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python vapt_complete.py <target_url>")
        print("Example: python vapt_complete.py https://example.com")
        sys.exit(1)
    
    target = sys.argv[1]
    result = perform_vapt_scan(target, owasp_enabled=True)
    
    if result['status'] == 'success':
        print(f"\n✅ Success! Report: {result['filename']}")
    else:
        print(f"\n❌ Error: {result['message']}")